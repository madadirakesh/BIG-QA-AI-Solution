import os

class ExcelExporter:
    @staticmethod
    def export_to_excel(locators: list, filepath: str) -> bool:
        """
        Exports a list of locator dictionaries to an Excel file.
        locators is a list of dicts: [{"name": "...", "type": "...", "action": "...", "value": "..."}, ...]
        """
        try:
            from openpyxl import Workbook, load_workbook
            
            if os.path.exists(filepath):
                try:
                    wb = load_workbook(filepath)
                    ws = wb.active
                except Exception as load_err:
                    print(f"Could not load existing Excel, creating new: {load_err}")
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Page Objects"
                    headers = ["Element Name", "Locator Type", "Action", "Locator Value"]
                    ws.append(headers)
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Page Objects"
                headers = ["Element Name", "Locator Type", "Action", "Locator Value"]
                ws.append(headers)

            # Data
            for loc in locators:
                row = [
                    loc.get("name") or loc.get("nameHint", ""),
                    loc.get("type", ""),
                    loc.get("action", ""),
                    loc.get("value", "")
                ]
                ws.append(row)

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(filepath)
            return True
        except Exception as e:
            print(f"Failed to export to Excel: {e}")
            return False
